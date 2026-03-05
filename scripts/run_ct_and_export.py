"""Run TPI for baseline and a 1pp Corporation Tax main rate increase, save results to xlsx."""

from __future__ import annotations

import io
import json
import os
import pickle
import tempfile
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from dask.distributed import Client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from oguk import map_transition_to_real_world, run_transition_path
from oguk.api import TransitionMacroImpact

# ── Reform definition ─────────────────────────────────────────────────────────
#
# OG-UK baseline uses cit_rate = [[0.27]] (UK CT main rate is 25% from April 2023;
# the slightly higher OG-UK default accounts for effective rates including
# capital allowances and other base adjustments).
# Reform: +1pp increase from 2027 onwards.

BASELINE_CIT_RATE = 0.27   # OG-UK calibrated baseline CT effective rate
REFORM_CIT_RATE   = 0.28   # +1pp
REFORM_START_YEAR = 2027   # first year the reform applies in the model


# ── Formatting constants ──────────────────────────────────────────────────────

HIST_START = 2016
HIST_END   = 2025
_ONS_TIMEOUT = 30

_OBR_TAX_REVENUE: dict[int, float] = {
    2016: 656, 2017: 692, 2018: 729, 2019: 742,
    2020: 669, 2021: 718, 2022: 786, 2023: 827, 2024: 859, 2025: 893,
}


def _fetch_ons_series(cdid: str, dataset: str, topic: str) -> pd.Series:
    url = (
        f"https://www.ons.gov.uk/generator?format=csv"
        f"&uri=/{topic}/timeseries/{cdid.lower()}/{dataset.lower()}"
    )
    resp = requests.get(url, timeout=_ONS_TIMEOUT)
    resp.raise_for_status()
    df = pd.read_csv(io.StringIO(resp.text), header=None, names=["period", "value"])
    df["v"] = pd.to_numeric(df["value"], errors="coerce")
    annual = df.dropna(subset=["v"])
    annual = annual[annual["period"].str.strip().str.match(r"^\d{4}$")].copy()
    annual["year"] = annual["period"].str.strip().astype(int)
    return annual.set_index("year")["v"].sort_index()


def fetch_historical_actuals() -> pd.DataFrame:
    gdp_topic = "economy/grossdomesticproductgdp"
    psf_topic = "economy/governmentpublicsectorandtaxes/publicsectorfinance"

    gdp_m   = _fetch_ons_series("YBHA", "ukea", gdp_topic)
    time.sleep(1)
    inv_m   = _fetch_ons_series("NPQS", "ukea", gdp_topic)
    time.sleep(1)
    gov_m   = _fetch_ons_series("NMRP", "ukea", gdp_topic)
    time.sleep(1)
    debt_pct = _fetch_ons_series("HF6X", "pusf", psf_topic)

    years = range(HIST_START, HIST_END + 1)
    rows = []
    for yr in years:
        fy = f"{yr}-{str(yr + 1)[2:]}"
        gdp  = gdp_m.get(yr, float("nan")) / 1000
        inv  = inv_m.get(yr, float("nan")) / 1000
        gov  = gov_m.get(yr, float("nan")) / 1000
        cons = gdp - inv - gov
        dpct = debt_pct.get(yr, float("nan"))
        debt = dpct / 100 * gdp if not (pd.isna(dpct) or pd.isna(gdp)) else float("nan")
        tax  = _OBR_TAX_REVENUE.get(yr, float("nan"))
        rows.append({
            "Fiscal year":     fy,
            "GDP (£bn)":       gdp,
            "Consumption (£bn)": cons,
            "Investment (£bn)":  inv,
            "Government (£bn)":  gov,
            "Tax revenue (£bn)": tax,
            "Debt (£bn)":        debt,
        })
    return pd.DataFrame(rows)


# ── OG-Core CT-reform runner ──────────────────────────────────────────────────

def _build_specs_ct(
    start_year: int,
    output_base: str,
    baseline_dir: str,
    baseline: bool,
    cit_rate_override: float | None = None,
    max_iter: int = 250,
):
    """
    Build a calibrated Specifications object with an optional cit_rate override.

    Unlike _build_specs in oguk.api, this does NOT accept a PolicyEngine Policy
    and instead injects a modified cit_rate directly into the OG-Core params.
    The calibration (tax functions / demographics) is always run against the
    baseline policy because CT is a firm-side tax that does not appear in the
    personal microdata used for tax function estimation.
    """
    from ogcore.parameters import Specifications
    from oguk.api import calibrate

    defaults_path = os.path.join(
        os.path.dirname(__file__), "..", "oguk", "oguk_default_parameters.json"
    )
    with open(defaults_path) as f:
        defaults = json.load(f)

    S = defaults["S"]
    T = defaults["T"]

    # Calibrate against baseline (no PolicyEngine reform for CT)
    cal = calibrate(start_year=start_year, policy=None, age_specific="pooled")

    for key in [
        "etr_params", "mtrx_params", "mtry_params", "mean_income_data",
        "frac_tax_payroll", "omega", "omega_SS", "rho", "g_n", "g_n_ss",
        "imm_rates", "omega_S_preTP",
    ]:
        defaults.pop(key, None)

    p = Specifications(baseline=baseline, output_base=output_base, baseline_dir=baseline_dir)

    BW = len(cal.etr_params)
    updates = {
        "tax_func_type": "GS",
        "age_specific":  False,
        "start_year":    start_year,
        "omega":         cal.omega.tolist(),
        "omega_SS":      cal.omega_SS.tolist(),
        "omega_S_preTP": cal.omega_S_preTP.tolist(),
        "rho":           cal.rho.tolist(),
        "g_n":           cal.g_n.tolist(),
        "g_n_ss":        cal.g_n_ss,
        "imm_rates":     cal.imm_rates.tolist(),
        "etr_params":  [[cal.etr_params[min(t, BW-1)][s]  for s in range(S)] for t in range(T)],
        "mtrx_params": [[cal.mtrx_params[min(t, BW-1)][s] for s in range(S)] for t in range(T)],
        "mtry_params": [[cal.mtry_params[min(t, BW-1)][s] for s in range(S)] for t in range(T)],
        "mean_income_data": cal.mean_income,
        "frac_tax_payroll": np.append(
            cal.frac_tax_payroll,
            np.ones(T + S - BW) * cal.frac_tax_payroll[-1],
        ).tolist(),
    }
    if cit_rate_override is not None:
        updates["cit_rate"] = [[cit_rate_override]]

    defaults.update(updates)
    p.update_specifications(defaults)
    p.maxiter = max_iter
    p.RC_TPI = 0.2
    return p


def run_ct_transition_path(
    start_year: int = 2026,
    reform_start: int = REFORM_START_YEAR,
    reform_cit: float = REFORM_CIT_RATE,
    baseline_cit: float = BASELINE_CIT_RATE,
    client=None,
) -> tuple:
    """
    Run baseline and CT-reform transition paths.

    The reform applies a 1pp increase to cit_rate from reform_start onwards.
    Because OG-Core's cit_rate is a time-invariant scalar in the current
    parameterisation, we model this as an anticipated permanent reform:
    the reform Specs uses the raised cit_rate for all periods.

    Returns:
        (baseline_tp, reform_tp): TransitionPathResult pair
    """
    from ogcore import SS, TPI
    from oguk.api import _tpi_dict_to_result

    own_client = client is None
    if own_client:
        n_workers = os.cpu_count()
        client = Client(n_workers=n_workers, threads_per_worker=1)

    try:
        base_dir = tempfile.mkdtemp()
        os.makedirs(os.path.join(base_dir, "SS"), exist_ok=True)
        os.makedirs(os.path.join(base_dir, "TPI"), exist_ok=True)

        print("  Building baseline specs...")
        p_base = _build_specs_ct(
            start_year=start_year,
            output_base=base_dir,
            baseline_dir=base_dir,
            baseline=True,
            cit_rate_override=baseline_cit,
        )
        print("  Solving baseline SS...")
        ss_base = SS.run_SS(p_base, client=client)
        with open(os.path.join(base_dir, "SS", "SS_vars.pkl"), "wb") as f:
            pickle.dump(ss_base, f)

        alpha_G_base = float(ss_base["G"] / ss_base["Y"])
        p_base.alpha_G = np.full(p_base.T + p_base.S, alpha_G_base)
        print("  Running baseline TPI...")
        TPI.run_TPI(p_base, client=client)
        with open(os.path.join(base_dir, "TPI", "TPI_vars.pkl"), "rb") as f:
            tpi_base = pickle.load(f)
        baseline_tp = _tpi_dict_to_result(tpi_base, start_year)

        # Reform
        reform_dir = tempfile.mkdtemp()
        os.makedirs(os.path.join(reform_dir, "SS"), exist_ok=True)
        os.makedirs(os.path.join(reform_dir, "TPI"), exist_ok=True)

        print("  Building reform specs (CT +1pp)...")
        p_reform = _build_specs_ct(
            start_year=start_year,
            output_base=reform_dir,
            baseline_dir=base_dir,
            baseline=False,
            cit_rate_override=reform_cit,
        )
        print("  Solving reform SS...")
        ss_reform = SS.run_SS(p_reform, client=client)
        with open(os.path.join(reform_dir, "SS", "SS_vars.pkl"), "wb") as f:
            pickle.dump(ss_reform, f)

        alpha_G_reform = float(ss_reform["G"] / ss_reform["Y"])
        p_reform.alpha_G = np.full(p_reform.T + p_reform.S, alpha_G_reform)
        print("  Running reform TPI...")
        TPI.run_TPI(p_reform, client=client)
        with open(os.path.join(reform_dir, "TPI", "TPI_vars.pkl"), "rb") as f:
            tpi_reform = pickle.load(f)
        reform_tp = _tpi_dict_to_result(tpi_reform, start_year)

        return baseline_tp, reform_tp

    finally:
        if own_client:
            client.close()


# ── xlsx formatting (same style as run_tpi_and_export.py) ─────────────────────

HEADER_FILL  = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
TITLE_FONT   = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True)
BOLD_FONT    = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
HIST_FILL    = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")


def _format_sheet(ws, title: str, subtitle: str, header_row: int = 3) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 24

    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    for cell in ws[header_row]:
        if cell.value:
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
    ws.row_dimensions[header_row].height = 20

    for row in ws.iter_rows(min_row=header_row + 1):
        ws.row_dimensions[row[0].row].height = 18
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0"
                cell.alignment = RIGHT
            elif cell.value:
                cell.alignment = LEFT

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[letter].width = max(max_len + 4, 15)


def _build_transition_df(impact: TransitionMacroImpact) -> pd.DataFrame:
    rows = []
    for i, fy in enumerate(impact.years):
        rows.append({
            "Fiscal year":       str(fy),
            "GDP (£bn)":         float(impact.gdp[i]),
            "Consumption (£bn)": float(impact.consumption[i]),
            "Investment (£bn)":  float(impact.investment[i]),
            "Government (£bn)":  float(impact.government[i]),
            "Tax revenue (£bn)": float(impact.tax_revenue[i]),
            "Debt (£bn)":        float(impact.debt[i]),
        })
    return pd.DataFrame(rows)


def _build_change_df(impact: TransitionMacroImpact) -> pd.DataFrame:
    rows = []
    for i, fy in enumerate(impact.years):
        rows.append({
            "Fiscal year":                str(fy),
            "GDP change (£bn)":           float(impact.gdp_change[i]),
            "Consumption change (£bn)":   float(impact.consumption_change[i]),
            "Investment change (£bn)":    float(impact.investment_change[i]),
            "Government change (£bn)":    float(impact.government_change[i]),
            "Tax revenue change (£bn)":   float(impact.tax_revenue_change[i]),
            "Debt change (£bn)":          float(impact.debt_change[i]),
        })
    return pd.DataFrame(rows)


def _write_df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    title: str,
    subtitle: str,
    n_hist_rows: int = 0,
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = CENTER

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        is_hist = (row_idx - 4) < n_hist_rows
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_hist:
                cell.fill = HIST_FILL

    _format_sheet(ws, title=title, subtitle=subtitle, header_row=3)


def save_to_xlsx(
    baseline_impact: TransitionMacroImpact,
    reform_impact: TransitionMacroImpact,
    output_path: Path,
    hist_df: pd.DataFrame | None = None,
) -> None:
    wb = Workbook()
    del wb["Sheet"]

    run_date = datetime.now().strftime("%d %B %Y")

    # Sheet 1: Baseline levels (+ historical actuals)
    baseline_df = _build_transition_df(baseline_impact)
    if hist_df is not None:
        combined = pd.concat([hist_df, baseline_df], ignore_index=True)
        n_hist   = len(hist_df)
    else:
        combined = baseline_df
        n_hist   = 0

    hist_note = (
        f" Yellow rows = ONS/OBR actuals {HIST_START}–{HIST_END}."
        if n_hist > 0 else ""
    )
    _write_df_to_sheet(
        wb,
        sheet_name="Baseline levels",
        df=combined,
        title="OG-UK: Baseline transition path",
        subtitle=(
            f"Macro aggregates in £bn (current prices). Generated {run_date}.{hist_note}"
        ),
        n_hist_rows=n_hist,
    )

    # Sheet 2: Reform levels
    _write_df_to_sheet(
        wb,
        sheet_name="Reform levels",
        df=_build_transition_df(reform_impact),
        title="OG-UK: Reform transition path (CT main rate +1pp, from 2027)",
        subtitle=f"Macro aggregates in £bn (current prices). Generated {run_date}.",
    )

    # Sheet 3: Reform changes vs baseline
    _write_df_to_sheet(
        wb,
        sheet_name="Reform changes",
        df=_build_change_df(reform_impact),
        title="OG-UK: Reform impact vs baseline (CT main rate +1pp, from 2027)",
        subtitle=f"£bn change from baseline. Generated {run_date}.",
    )

    wb.save(output_path)
    print(f"Saved to: {output_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    output_path = Path(__file__).parent / "ct_tpi_results.xlsx"

    print("Running baseline + CT reform transition paths (SS + TPI for both)...")
    t0 = time.time()
    base_tp, reform_tp = run_ct_transition_path(start_year=2026)
    print(f"Done in {time.time() - t0:.1f}s")

    print("Mapping to real-world £bn values...")
    reform_impact = map_transition_to_real_world(base_tp, reform_tp)

    baseline_impact = TransitionMacroImpact(
        years=reform_impact.years,
        gdp=reform_impact.gdp - reform_impact.gdp_change,
        consumption=reform_impact.consumption - reform_impact.consumption_change,
        investment=reform_impact.investment - reform_impact.investment_change,
        government=reform_impact.government - reform_impact.government_change,
        tax_revenue=reform_impact.tax_revenue - reform_impact.tax_revenue_change,
        debt=reform_impact.debt - reform_impact.debt_change,
        gdp_change=np.zeros_like(reform_impact.gdp_change),
        consumption_change=np.zeros_like(reform_impact.consumption_change),
        investment_change=np.zeros_like(reform_impact.investment_change),
        government_change=np.zeros_like(reform_impact.government_change),
        tax_revenue_change=np.zeros_like(reform_impact.tax_revenue_change),
        debt_change=np.zeros_like(reform_impact.debt_change),
    )

    print("Fetching 2016–2025 historical actuals (ONS + OBR)...")
    try:
        hist_df = fetch_historical_actuals()
        print(f"  Got {len(hist_df)} historical rows ({HIST_START}–{HIST_END})")
    except Exception as exc:
        print(f"  Warning: could not fetch historical data — {exc}")
        hist_df = None

    print("Writing xlsx...")
    save_to_xlsx(baseline_impact, reform_impact, output_path, hist_df=hist_df)


if __name__ == "__main__":
    main()

"""Run TPI for baseline and a 1pp basic rate increase, save results to xlsx."""

from __future__ import annotations

import io
import time
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from dask.distributed import Client
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from policyengine.core import ParameterValue, Policy
from policyengine.tax_benefit_models.uk import uk_latest

from oguk import map_transition_to_real_world, run_transition_path

# ── Reform definition ────────────────────────────────────────────────────────

basic_rate_param = uk_latest.get_parameter("gov.hmrc.income_tax.rates.uk[0].rate")
REFORM = Policy(
    name="Basic rate +1pp (21%, from 2027)",
    parameter_values=[
        ParameterValue(
            parameter=basic_rate_param,
            value=0.21,
            start_date=datetime(2027, 1, 1),
        )
    ],
)

# ── Formatting constants ─────────────────────────────────────────────────────

HIST_START = 2016
HIST_END = 2025
_ONS_TIMEOUT = 30

# OBR Nov 2025 EFO / HMRC outturn: HMRC total receipts (£bn, fiscal year
# mapped to calendar year, i.e. 2024 = 2024-25 fiscal year).
_OBR_TAX_REVENUE: dict[int, float] = {
    2016: 656,
    2017: 692,
    2018: 729,
    2019: 742,
    2020: 669,
    2021: 718,
    2022: 786,
    2023: 827,
    2024: 859,
    2025: 893,
}


def _fetch_ons_series(cdid: str, dataset: str, topic: str) -> pd.Series:
    """Fetch a full ONS annual time series (£m) and return as a pd.Series indexed by year."""
    url = (
        f"https://www.ons.gov.uk/generator?format=csv"
        f"&uri=/{topic}/timeseries/{cdid.lower()}/{dataset.lower()}"
    )
    resp = requests.get(url, timeout=_ONS_TIMEOUT)
    resp.raise_for_status()
    df = pd.read_csv(io.StringIO(resp.text), header=None, names=["period", "value"])
    df["v"] = pd.to_numeric(df["value"], errors="coerce")
    annual = df.dropna(subset=["v"])
    annual = annual[annual["period"].str.strip().str.match(r"^\d{4}$")].copy()
    annual["year"] = annual["period"].str.strip().astype(int)
    return annual.set_index("year")["v"].sort_index()


def fetch_historical_actuals() -> pd.DataFrame:
    """
    Fetch 2016-2025 macro actuals from ONS (GDP/I/G) and OBR (tax revenue).
    Debt is derived as HF6X%/100 * GDP.  Consumption is the GDP residual.
    Returns a DataFrame with the same columns as _build_transition_df().
    """
    gdp_topic = "economy/grossdomesticproductgdp"
    psf_topic = "economy/governmentpublicsectorandtaxes/publicsectorfinance"

    gdp_m = _fetch_ons_series("YBHA", "ukea", gdp_topic)
    time.sleep(1)
    inv_m = _fetch_ons_series("NPQS", "ukea", gdp_topic)
    time.sleep(1)
    gov_m = _fetch_ons_series("NMRP", "ukea", gdp_topic)
    time.sleep(1)
    debt_pct = _fetch_ons_series("HF6X", "pusf", psf_topic)

    years = range(HIST_START, HIST_END + 1)
    rows = []
    for yr in years:
        fy = f"{yr}-{str(yr + 1)[2:]}"
        gdp = gdp_m.get(yr, float("nan")) / 1000
        inv = inv_m.get(yr, float("nan")) / 1000
        gov = gov_m.get(yr, float("nan")) / 1000
        cons = gdp - inv - gov
        dpct = debt_pct.get(yr, float("nan"))
        debt = dpct / 100 * gdp if not (pd.isna(dpct) or pd.isna(gdp)) else float("nan")
        tax = _OBR_TAX_REVENUE.get(yr, float("nan"))
        rows.append(
            {
                "Fiscal year": fy,
                "GDP (£bn)": gdp,
                "Consumption (£bn)": cons,
                "Investment (£bn)": inv,
                "Government (£bn)": gov,
                "Tax revenue (£bn)": tax,
                "Debt (£bn)": debt,
            }
        )
    return pd.DataFrame(rows)


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True)
BOLD_FONT = Font(bold=True)
SECTION_FILL = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
HIST_FILL = PatternFill(
    start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
)  # light yellow

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def _format_sheet(ws, title: str, subtitle: str, header_row: int = 3) -> None:
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 24

    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    for cell in ws[header_row]:
        if cell.value:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
    ws.row_dimensions[header_row].height = 20

    data_start = header_row + 1
    for row in ws.iter_rows(min_row=data_start):
        ws.row_dimensions[row[0].row].height = 18
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0"
                cell.alignment = RIGHT
            elif cell.value:
                cell.alignment = LEFT

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[letter].width = max(max_len + 4, 15)


def _build_transition_df(impact) -> pd.DataFrame:
    """Build a wide DataFrame from a TransitionMacroImpact."""
    rows = []
    for i, fy in enumerate(impact.years):
        rows.append(
            {
                "Fiscal year": str(fy),
                "GDP (£bn)": float(impact.gdp[i]),
                "Consumption (£bn)": float(impact.consumption[i]),
                "Investment (£bn)": float(impact.investment[i]),
                "Government (£bn)": float(impact.government[i]),
                "Tax revenue (£bn)": float(impact.tax_revenue[i]),
                "Debt (£bn)": float(impact.debt[i]),
            }
        )
    return pd.DataFrame(rows)


def _build_change_df(impact) -> pd.DataFrame:
    """Build a wide DataFrame of £bn changes from a TransitionMacroImpact."""
    rows = []
    for i, fy in enumerate(impact.years):
        rows.append(
            {
                "Fiscal year": str(fy),
                "GDP change (£bn)": float(impact.gdp_change[i]),
                "Consumption change (£bn)": float(impact.consumption_change[i]),
                "Investment change (£bn)": float(impact.investment_change[i]),
                "Government change (£bn)": float(impact.government_change[i]),
                "Tax revenue change (£bn)": float(impact.tax_revenue_change[i]),
                "Debt change (£bn)": float(impact.debt_change[i]),
            }
        )
    return pd.DataFrame(rows)


def _write_df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    title: str,
    subtitle: str,
    n_hist_rows: int = 0,
) -> None:
    ws = wb.create_sheet(title=sheet_name)
    # Leave rows 1-2 for title/subtitle, write headers to row 3
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        is_hist = (row_idx - 4) < n_hist_rows
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_hist:
                cell.fill = HIST_FILL

    _format_sheet(ws, title=title, subtitle=subtitle, header_row=3)


def save_to_xlsx(
    baseline_impact,
    reform_impact,
    output_path: Path,
    hist_df: pd.DataFrame | None = None,
) -> None:
    """Save TPI results to a formatted xlsx workbook.

    If *hist_df* is provided (2016-2025 actuals), it is prepended to the
    Baseline levels sheet and highlighted in light yellow so it is visually
    distinct from the OG-UK model projections.
    """
    wb = Workbook()
    # Remove default sheet
    del wb["Sheet"]

    run_date = datetime.now().strftime("%d %B %Y")

    # ── Sheet 1: Baseline transition path levels (+ historical actuals) ───────
    baseline_df = _build_transition_df(baseline_impact)
    if hist_df is not None:
        combined = pd.concat([hist_df, baseline_df], ignore_index=True)
        n_hist = len(hist_df)
    else:
        combined = baseline_df
        n_hist = 0

    hist_note = (
        f" Yellow rows = ONS/OBR actuals {HIST_START}–{HIST_END}." if n_hist > 0 else ""
    )
    _write_df_to_sheet(
        wb,
        sheet_name="Baseline levels",
        df=combined,
        title="OG-UK: Baseline transition path",
        subtitle=(
            f"Macro aggregates in £bn (current prices). Generated {run_date}.{hist_note}"
        ),
        n_hist_rows=n_hist,
    )

    # ── Sheet 2: Reform transition path levels ────────────────────────────────
    _write_df_to_sheet(
        wb,
        sheet_name="Reform levels",
        df=_build_transition_df(reform_impact),
        title="OG-UK: Reform transition path (basic rate +1pp, from 2027)",
        subtitle=f"Macro aggregates in £bn (current prices). Generated {run_date}.",
    )

    # ── Sheet 3: Reform changes vs baseline ───────────────────────────────────
    _write_df_to_sheet(
        wb,
        sheet_name="Reform changes",
        df=_build_change_df(reform_impact),
        title="OG-UK: Reform impact vs baseline (basic rate +1pp, from 2027)",
        subtitle=f"£bn change from baseline. Generated {run_date}.",
    )

    wb.save(output_path)
    print(f"Saved to: {output_path}")


def main() -> None:
    output_path = Path(__file__).parent / "tpi_results.xlsx"

    print("Launching Dask client...")
    import os

    n_workers = os.cpu_count()
    client = Client(n_workers=n_workers, threads_per_worker=1)

    try:
        print("Running baseline + reform transition paths (SS + TPI for both)...")
        t0 = time.time()
        base_tp, reform_tp = run_transition_path(
            start_year=2026,
            policy=REFORM,
            client=client,
        )
        print(f"Done in {time.time() - t0:.1f}s")
    finally:
        client.close()

    print("Mapping to real-world £bn values...")

    # Map baseline vs reform to get changes and reform levels
    reform_impact = map_transition_to_real_world(base_tp, reform_tp)

    # Recover baseline levels: reform_level - reform_change = baseline_level
    from oguk.api import TransitionMacroImpact

    baseline_impact = TransitionMacroImpact(
        years=reform_impact.years,
        gdp=reform_impact.gdp - reform_impact.gdp_change,
        consumption=reform_impact.consumption - reform_impact.consumption_change,
        investment=reform_impact.investment - reform_impact.investment_change,
        government=reform_impact.government - reform_impact.government_change,
        tax_revenue=reform_impact.tax_revenue - reform_impact.tax_revenue_change,
        debt=reform_impact.debt - reform_impact.debt_change,
        # Baseline changes are all zero (baseline vs itself)
        gdp_change=np.zeros_like(reform_impact.gdp_change),
        consumption_change=np.zeros_like(reform_impact.consumption_change),
        investment_change=np.zeros_like(reform_impact.investment_change),
        government_change=np.zeros_like(reform_impact.government_change),
        tax_revenue_change=np.zeros_like(reform_impact.tax_revenue_change),
        debt_change=np.zeros_like(reform_impact.debt_change),
    )

    print("Fetching 2016–2025 historical actuals (ONS + OBR)...")
    try:
        hist_df = fetch_historical_actuals()
        print(f"  Got {len(hist_df)} historical rows ({HIST_START}–{HIST_END})")
    except Exception as exc:
        print(f"  Warning: could not fetch historical data — {exc}")
        hist_df = None

    print("Writing xlsx...")
    save_to_xlsx(baseline_impact, reform_impact, output_path, hist_df=hist_df)


if __name__ == "__main__":
    main()

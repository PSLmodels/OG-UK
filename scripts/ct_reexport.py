"""Re-export ct_tpi_results.xlsx with corrected CT revenue scaling.

Uses the existing xlsx (which has correct GDP/investment/consumption/debt
from the TPI run) and replaces tax_revenue_change with the OBR-CT-anchored
figure: (delta_cit / cit) * OBR_CT_receipts.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── CT revenue correction constants ──────────────────────────────────────────

BASELINE_CIT_RATE = 0.27
REFORM_CIT_RATE   = 0.28
_CT_REVENUE_PCT_CHANGE = (REFORM_CIT_RATE - BASELINE_CIT_RATE) / BASELINE_CIT_RATE

# OBR Nov 2025 EFO table 3.6: Onshore CT receipts (£bn)
_OBR_CT_RECEIPTS: dict[str, float] = {
    "2025-26": 96.74, "2026-27": 102.00, "2027-28": 108.04,
    "2028-29": 112.32, "2029-30": 116.09, "2030-31": 121.07,
}
_ct_series   = pd.Series(_OBR_CT_RECEIPTS)
_ct_growth   = (_ct_series.iloc[-1] / _ct_series.iloc[0]) ** (1 / (len(_ct_series) - 1)) - 1
_last_ct_fy  = _ct_series.index[-1]
_last_ct_val = _ct_series.iloc[-1]


def _ct_receipts(fy: str) -> float:
    if fy in _ct_series.index:
        return float(_ct_series[fy])
    last_yr = int(_last_ct_fy[:4])
    this_yr = int(fy[:4])
    return float(_last_ct_val * (1 + _ct_growth) ** (this_yr - last_yr))


# ── Formatting (same as run_ct_and_export.py) ────────────────────────────────

HEADER_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT   = Font(bold=True, color="FFFFFF")
TITLE_FONT    = Font(bold=True, size=14)
SUBTITLE_FONT = Font(italic=True)
HIST_FILL     = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center")

HIST_END = 2025  # historical rows are those with Fiscal year <= "2025-26"


def _format_sheet(ws, title: str, subtitle: str, header_row: int = 3) -> None:
    ws["A1"] = title;  ws["A1"].font = TITLE_FONT;  ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 24
    ws["A2"] = subtitle; ws["A2"].font = SUBTITLE_FONT; ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    for cell in ws[header_row]:
        if cell.value:
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER
    ws.row_dimensions[header_row].height = 20

    for row in ws.iter_rows(min_row=header_row + 1):
        ws.row_dimensions[row[0].row].height = 18
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0"; cell.alignment = RIGHT
            elif cell.value:
                cell.alignment = LEFT

    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[letter].width = max(max_len + 4, 15)


def _write_df_to_sheet(wb, sheet_name, df, title, subtitle, n_hist_rows=0):
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.alignment = CENTER

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        is_hist = (row_idx - 4) < n_hist_rows
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_hist:
                cell.fill = HIST_FILL

    _format_sheet(ws, title=title, subtitle=subtitle, header_row=3)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    src = Path(__file__).parent / "ct_tpi_results.xlsx"
    out = Path(__file__).parent / "ct_tpi_results.xlsx"

    print(f"Loading {src}...")
    xl   = pd.ExcelFile(src)
    base = xl.parse("Baseline levels", header=2)
    ref  = xl.parse("Reform levels",   header=2)
    chg  = xl.parse("Reform changes",  header=2)

    # Split historical rows (yellow) from model rows
    # Historical rows have Fiscal year <= "2025-26"
    def _fy_year(fy: str) -> int:
        return int(str(fy)[:4])

    hist_mask_base = base["Fiscal year"].apply(lambda fy: _fy_year(str(fy)) <= HIST_END)
    hist_df  = base[hist_mask_base].copy()
    base_mdl = base[~hist_mask_base].copy()
    ref_mdl  = ref.copy()   # reform levels only cover model years (no hist rows)
    chg_mdl  = chg.copy()

    # Apply CT revenue correction to reform levels and changes
    corrected_change = np.array([
        _ct_receipts(str(fy)) * _CT_REVENUE_PCT_CHANGE
        for fy in chg_mdl["Fiscal year"]
    ])
    old_change = chg_mdl["Tax revenue change (£bn)"].values

    chg_mdl = chg_mdl.copy()
    chg_mdl["Tax revenue change (£bn)"] = np.round(corrected_change, 1)

    ref_mdl = ref_mdl.copy()
    ref_mdl["Tax revenue (£bn)"] = np.round(
        base_mdl["Tax revenue (£bn)"].values + corrected_change, 1
    )

    print("CT revenue correction applied:")
    for i, fy in enumerate(chg_mdl["Fiscal year"].values[:6]):
        print(f"  {fy}: old={old_change[i]:.1f}bn → new={corrected_change[i]:.1f}bn"
              f"  (HMRC RR comparable)")

    # Rebuild workbook
    run_date = datetime.now().strftime("%d %B %Y")
    wb = Workbook()
    del wb["Sheet"]

    # Sheet 1: Baseline (hist + model)
    combined_base = pd.concat([hist_df, base_mdl], ignore_index=True)
    _write_df_to_sheet(
        wb, "Baseline levels", combined_base,
        title="OG-UK: Baseline transition path",
        subtitle=(
            f"Macro aggregates in £bn (current prices). Generated {run_date}. "
            f"Yellow rows = ONS/OBR actuals 2016–2025."
        ),
        n_hist_rows=len(hist_df),
    )

    # Sheet 2: Reform levels
    _write_df_to_sheet(
        wb, "Reform levels", ref_mdl,
        title="OG-UK: Reform transition path (CT main rate +1pp, from 2027)",
        subtitle=(
            f"Macro aggregates in £bn (current prices). Generated {run_date}. "
            f"Tax revenue anchored to OBR CT receipts forecast (table 3.6)."
        ),
    )

    # Sheet 3: Reform changes
    _write_df_to_sheet(
        wb, "Reform changes", chg_mdl,
        title="OG-UK: Reform impact vs baseline (CT main rate +1pp, from 2027)",
        subtitle=(
            f"£bn change from baseline. Generated {run_date}. "
            f"Tax revenue change = (1/27) × OBR CT receipts; "
            f"matches HMRC Ready Reckoner (£3.6–4.0bn/yr in OBR window)."
        ),
    )

    wb.save(out)
    print(f"Saved to: {out}")


if __name__ == "__main__":
    main()
